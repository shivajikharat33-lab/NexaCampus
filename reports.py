from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
import io, csv

from database import complaints
from auth_middleware import require_admin

router = APIRouter(prefix="/reports", tags=["Reports"])

FIELDS = [
    "complaint_id", "student_id", "student_name", "department", "year",
    "category", "description", "location", "priority", "status",
    "assigned_to", "upvotes", "resolution_time_hours", "created_at", "updated_at"
]


def _get_complaints(query: dict = {}) -> list:
    return list(complaints.find(query, {"_id": 0}))


def _fmt(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    return str(val) if val is not None else ""


# ── CSV Export ────────────────────────────────────────────────────
@router.get("/export/csv")
def export_csv(
    status: str = None,
    department: str = None,
    current_user: dict = Depends(require_admin)
):
    query = {}
    if status:     query["status"] = status
    if department: query["assigned_to"] = department

    data = _get_complaints(query)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=FIELDS, extrasaction="ignore")
    writer.writeheader()
    for row in data:
        writer.writerow({f: _fmt(row.get(f)) for f in FIELDS})

    output.seek(0)
    filename = f"complaints_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── Excel Export ──────────────────────────────────────────────────
@router.get("/export/excel")
def export_excel(
    status: str = None,
    department: str = None,
    current_user: dict = Depends(require_admin)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    query = {}
    if status:     query["status"] = status
    if department: query["assigned_to"] = department

    data = _get_complaints(query)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Complaints"

    # Header styling
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")

    for col, field in enumerate(FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=field.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Status color map
    status_colors = {
        "Pending":     "FFF3CD",
        "In Progress": "CCE5FF",
        "Resolved":    "D4EDDA",
        "Rejected":    "F8D7DA",
        "Reopened":    "E2D9F3"
    }

    for row_num, item in enumerate(data, 2):
        for col, field in enumerate(FIELDS, 1):
            ws.cell(row=row_num, column=col, value=_fmt(item.get(field)))
        # Color row by status
        status = item.get("status", "")
        color = status_colors.get(status)
        if color:
            fill = PatternFill("solid", fgColor=color)
            for col in range(1, len(FIELDS) + 1):
                ws.cell(row=row_num, column=col).fill = fill

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"complaints_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── PDF Export ────────────────────────────────────────────────────
@router.get("/export/pdf")
def export_pdf(
    status: str = None,
    department: str = None,
    current_user: dict = Depends(require_admin)
):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="reportlab not installed. Run: pip install reportlab")

    query = {}
    if status:     query["status"] = status
    if department: query["assigned_to"] = department

    data = _get_complaints(query)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=10*mm, rightMargin=10*mm)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(
        f"Smart College Civic Detector — Complaints Report",
        styles["Title"]
    ))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Total: {len(data)}",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 6*mm))

    # Table
    short_fields = ["complaint_id", "student_name", "category", "location", "priority", "status", "assigned_to", "created_at"]
    headers = [f.replace("_", " ").title() for f in short_fields]
    table_data = [headers]
    for item in data:
        table_data.append([_fmt(item.get(f)) for f in short_fields])

    col_widths = [35*mm, 30*mm, 28*mm, 30*mm, 20*mm, 22*mm, 28*mm, 28*mm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)

    output.seek(0)
    filename = f"complaints_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── Monthly Report ────────────────────────────────────────────────
@router.get("/monthly")
def monthly_report(year: int = None, month: int = None, current_user: dict = Depends(require_admin)):
    now = datetime.now(timezone.utc)
    y   = year  or now.year
    m   = month or now.month
    start = datetime(y, m, 1, tzinfo=timezone.utc)
    end   = datetime(y, m + 1, 1, tzinfo=timezone.utc) if m < 12 else datetime(y + 1, 1, 1, tzinfo=timezone.utc)

    q = {"created_at": {"$gte": start, "$lt": end}}
    return {
        "period": f"{y}-{m:02d}",
        "total":       complaints.count_documents(q),
        "resolved":    complaints.count_documents({**q, "status": "Resolved"}),
        "pending":     complaints.count_documents({**q, "status": "Pending"}),
        "rejected":    complaints.count_documents({**q, "status": "Rejected"}),
        "emergency":   complaints.count_documents({**q, "priority": "Emergency"}),
    }


# ── Department Report ─────────────────────────────────────────────
@router.get("/department")
def department_report(department: str, current_user: dict = Depends(require_admin)):
    q = {"assigned_to": department}
    total    = complaints.count_documents(q)
    resolved = complaints.count_documents({**q, "status": "Resolved"})
    pending  = complaints.count_documents({**q, "status": "Pending"})
    avg_q = list(complaints.aggregate([
        {"$match": {**q, "resolution_time_hours": {"$ne": None}}},
        {"$group": {"_id": None, "avg": {"$avg": "$resolution_time_hours"}}}
    ]))
    return {
        "department": department,
        "total": total,
        "resolved": resolved,
        "pending": pending,
        "resolution_rate_%": round(resolved / total * 100, 1) if total else 0,
        "avg_resolution_hours": round(avg_q[0]["avg"], 2) if avg_q else 0
    }